from __future__ import annotations

from datetime import datetime
import random
import re
from uuid import uuid4

from flask import (
    Blueprint,
    flash,
    redirect,
    render_template,
    request,
    session,
    url_for,
)
from flask_login import current_user, login_required, login_user, logout_user
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from urllib.parse import urljoin, urlparse

from .. import db
from ..i18n import DEFAULT_LANGUAGE, get_language_choices
from ..models import (
    Admin,
    Appointment,
    AvailabilitySlot,
    Coach,
    ExamRule,
    MockExamPaper,
    MockExamPaperQuestion,
    MockExamSummary,
    Question,
    Student,
)
from ..services import StateSwitchError, switch_student_state

coach_bp = Blueprint("coach", __name__, url_prefix="/coach")

STATE_CHOICES: list[str] = [
    "ACT",
    "NSW",
    "NT",
    "QLD",
    "SA",
    "TAS",
    "VIC",
    "WA",
]

LANGUAGE_CODES: list[str] = [choice["code"] for choice in get_language_choices()]
VALID_OPTIONS = {"A", "B", "C", "D"}

def _calling_code_entry(
    code: str,
    iso: str,
    flag: str,
    english_name: str,
    chinese_name: str,
) -> dict[str, str | dict[str, str]]:
    digits = code.lstrip("+")
    search_terms = {
        english_name.lower(),
        chinese_name.lower(),
        english_name.replace(" ", "").lower(),
        chinese_name.replace(" ", "").lower(),
        iso.lower(),
        code.lower(),
        digits,
    }
    return {
        "code": code,
        "iso": iso,
        "flag": flag,
        "names": {"ENGLISH": english_name, "CHINESE": chinese_name},
        "search_terms": " ".join(sorted(search_terms)),
    }


COUNTRY_CALLING_CODES: list[dict[str, str | dict[str, str]]] = [
    _calling_code_entry("+61", "AU", "🇦🇺", "Australia", "澳大利亚"),
    _calling_code_entry("+86", "CN", "🇨🇳", "China mainland", "中国大陆"),
    _calling_code_entry("+852", "HK", "🇭🇰", "Hong Kong", "中国香港"),
    _calling_code_entry("+886", "TW", "🇹🇼", "Taiwan", "中国台湾"),
    _calling_code_entry("+65", "SG", "🇸🇬", "Singapore", "新加坡"),
    _calling_code_entry("+81", "JP", "🇯🇵", "Japan", "日本"),
    _calling_code_entry("+82", "KR", "🇰🇷", "South Korea", "韩国"),
    _calling_code_entry("+60", "MY", "🇲🇾", "Malaysia", "马来西亚"),
    _calling_code_entry("+64", "NZ", "🇳🇿", "New Zealand", "新西兰"),
    _calling_code_entry("+44", "GB", "🇬🇧", "United Kingdom", "英国"),
    _calling_code_entry("+1", "US", "🇺🇸", "United States & Canada", "美国 / 加拿大"),
    _calling_code_entry("+62", "ID", "🇮🇩", "Indonesia", "印度尼西亚"),
    _calling_code_entry("+63", "PH", "🇵🇭", "Philippines", "菲律宾"),
    _calling_code_entry("+66", "TH", "🇹🇭", "Thailand", "泰国"),
    _calling_code_entry("+84", "VN", "🇻🇳", "Vietnam", "越南"),
    _calling_code_entry("+91", "IN", "🇮🇳", "India", "印度"),
]

LANGUAGE_DEFAULT_CALLING_CODES: dict[str, str] = {
    "ENGLISH": "+61",
    "CHINESE": "+86",
}

DEFAULT_CALLING_CODE = LANGUAGE_DEFAULT_CALLING_CODES[DEFAULT_LANGUAGE]


def _normalize_mobile_number(raw: str) -> str:
    return "".join(ch for ch in (raw or "") if ch.isdigit())


def _calling_code_digits(raw_code: str | None) -> str:
    return "".join(ch for ch in (raw_code or "") if ch.isdigit())


def _strip_trunk_prefix(local_digits: str) -> str:
    if local_digits.startswith("0"):
        stripped = local_digits.lstrip("0")
        if stripped:
            return stripped
    return local_digits


def _combine_calling_code_and_local_number(
    raw_code: str, raw_local: str
) -> str:
    local_digits = _normalize_mobile_number(raw_local)
    if not local_digits:
        return ""
    code_digits = _calling_code_digits(raw_code)
    if code_digits:
        local_digits = _strip_trunk_prefix(local_digits)
        if not local_digits:
            return ""
        return f"{code_digits}{local_digits}"
    return local_digits


def _normalise_mobile_with_default(raw_input: str) -> str:
    digits = _normalize_mobile_number(raw_input)
    if not digits:
        return ""

    for entry in COUNTRY_CALLING_CODES:
        code_digits = _calling_code_digits(str(entry["code"]))
        if code_digits and digits.startswith(code_digits):
            return digits

    default_digits = _calling_code_digits(DEFAULT_CALLING_CODE)
    if not default_digits:
        return digits

    stripped_local = _strip_trunk_prefix(digits)
    if not stripped_local:
        return ""

    return f"{default_digits}{stripped_local}"


def _candidate_mobile_numbers(
    raw_input: str, selected_code: str | None = None
) -> list[str]:
    normalized = _normalize_mobile_number(raw_input)
    if not normalized:
        return []

    variants: list[str] = []
    seen: set[str] = set()

    def _add(value: str) -> None:
        if value and value not in seen:
            variants.append(value)
            seen.add(value)

    _add(normalized)

    has_known_prefix = False
    for entry in COUNTRY_CALLING_CODES:
        code_digits = _calling_code_digits(str(entry["code"]))
        if code_digits and normalized.startswith(code_digits):
            has_known_prefix = True
            remainder = normalized[len(code_digits) :]
            if remainder:
                normalized_remainder = _strip_trunk_prefix(remainder)
                _add(f"{code_digits}{normalized_remainder}")
            break

    trimmed = _strip_trunk_prefix(normalized)
    if trimmed and trimmed != normalized:
        _add(trimmed)

    if has_known_prefix:
        return variants

    basis = trimmed or normalized
    if not basis:
        return variants

    if selected_code:
        code_digits = _calling_code_digits(selected_code)
        if code_digits:
            _add(f"{code_digits}{basis}")
    else:
        for entry in COUNTRY_CALLING_CODES:
            code_digits = _calling_code_digits(str(entry["code"]))
            if code_digits:
                _add(f"{code_digits}{basis}")

    return variants


def _locate_account_by_mobile(model, raw_input: str, selected_code: str | None = None):
    variants = _candidate_mobile_numbers(raw_input, selected_code)
    if not variants:
        return None, False

    normalized_column = _normalized_mobile_expression(model.mobile_number)
    matches = (
        model.query.filter(normalized_column.in_(variants))
        .order_by(model.id.asc())
        .all()
    )
    if not matches:
        return None, False

    normalized_to_records: dict[str, list] = {}
    for record in matches:
        normalized_value = _normalize_mobile_number(record.mobile_number)
        normalized_to_records.setdefault(normalized_value, []).append(record)

    if len(normalized_to_records) != 1:
        return None, True

    records = next(iter(normalized_to_records.values()))
    if len(records) > 1:
        return None, True

    normalized_value = _normalize_mobile_number(records[0].mobile_number)
    for variant in variants:
        if variant == normalized_value:
            return records[0], False

    return records[0], False


def _normalized_mobile_expression(column):
    sanitized = column
    for character in (" ", "-", "(", ")", "+"):
        sanitized = db.func.replace(sanitized, character, "")
    return sanitized


def _parse_vehicle_type(value: str | None) -> str | None:
    allowed = {"AT", "MT"}
    if value is None:
        return None
    cleaned = value.strip().upper()
    if cleaned not in allowed:
        return None
    return cleaned


def _extract_vehicle_type_from_form() -> str | None:
    parsed = _parse_vehicle_type(request.form.get("vehicle_type"))
    if parsed:
        return parsed
    legacy_values = request.form.getlist("vehicle_types")
    for entry in legacy_values:
        parsed = _parse_vehicle_type(entry)
        if parsed:
            return parsed
    return None


def _normalize_mobile(raw_value: str) -> str:
    digits = "".join(ch for ch in raw_value if ch.isdigit())
    if digits:
        return digits
    return raw_value.strip()


def _require_admin_access():
    if not current_user.is_admin:
        flash("Only administrators may access personnel management.", "danger")
        return redirect(url_for("coach.dashboard"))
    return None


def _question_base_query(state: str | None) -> "db.Query":
    query = Question.query
    if state:
        query = query.filter(or_(Question.state_scope == "ALL", Question.state_scope == state))
    return query.order_by(Question.topic.asc(), Question.qid.asc())


def _default_question_state(raw_state: str | None) -> str:
    if raw_state:
        return raw_state
    if current_user.is_admin:
        return "ALL"
    return current_user.state


@coach_bp.before_app_request
def _restrict_student_portal_access():
    if request.blueprint != "coach":
        return None
    if request.endpoint in {"coach.login", "coach.register_student", "coach.logout"}:
        return None
    if request.endpoint is None:
        return None
    if not current_user.is_authenticated:
        return None
    if getattr(current_user, "is_student", False):
        flash("Student accounts should use the learner portal.", "warning")
        return redirect(url_for("student.dashboard"))
    return None

def _is_safe_redirect_target(target: str | None) -> bool:
    if not target:
        return False
    host_url = request.host_url
    redirect_url = urljoin(host_url, target)
    host_parts = urlparse(host_url)
    redirect_parts = urlparse(redirect_url)
    return host_parts.scheme == redirect_parts.scheme and host_parts.netloc == redirect_parts.netloc


@coach_bp.route("/login", methods=["GET", "POST"])
def login():
    def _render_form():
        return render_template("coach/login.html")

    if request.method == "POST":
        mobile_input = (request.form.get("mobile_number") or "").strip()
        country_code_input = DEFAULT_CALLING_CODE
        password = request.form.get("password", "")
        next_url = request.args.get("next")
        if not _is_safe_redirect_target(next_url):
            next_url = None

        coach = None
        ambiguous_mobile = False
        if mobile_input:
            coach, ambiguous_mobile = _locate_account_by_mobile(
                Coach, mobile_input, country_code_input
            )

        if not ambiguous_mobile and coach and coach.check_password(password):
            login_user(coach)
            flash("Welcome back!", "success")
            return redirect(next_url or url_for("coach.dashboard"))

        student = None
        if not ambiguous_mobile and coach is None and mobile_input:
            student, ambiguous_mobile = _locate_account_by_mobile(
                Student, mobile_input, country_code_input
            )

        if not ambiguous_mobile and student and student.check_password(password):
            login_user(student)
            flash("Welcome back!", "success")
            return redirect(next_url or url_for("student.dashboard"))

        if ambiguous_mobile:
            flash(
                "Multiple accounts match that mobile number. Please include your country calling code.",
                "danger",
            )
        else:
            flash("Invalid mobile number or password", "danger")

        return _render_form()

    return _render_form()


@coach_bp.route("/register", methods=["GET", "POST"])
def register_student():
    def _render_form():
        return render_template(
            "coach/register_student.html",
            state_choices=STATE_CHOICES,
        )

    if request.method == "GET":
        return _render_form()

    name = (request.form.get("student_name") or "").strip()
    mobile_input = (request.form.get("student_mobile_number") or "").strip()
    mobile_number = _normalise_mobile_with_default(mobile_input)
    if not mobile_number:
        flash("Please enter a valid mobile number.", "danger")
        return _render_form()
    email = (request.form.get("student_email") or "").strip() or None
    password = request.form.get("student_password", "")
    confirm_password = request.form.get("student_confirm_password", "")
    state_choice = (request.form.get("student_state") or "").strip().upper()
    preferred_language = (
        (request.form.get("student_preferred_language") or "ENGLISH").strip().upper()
    )

    if not name or not mobile_number or not password:
        flash("Name, mobile number, and password are required to register.", "danger")
        return _render_form()

    if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        flash("Please enter a valid email address.", "danger")
        return _render_form()

    if password != confirm_password:
        flash("Passwords do not match.", "danger")
        return _render_form()

    if state_choice not in STATE_CHOICES:
        flash("Please select a valid state or territory.", "danger")
        return _render_form()

    if preferred_language not in LANGUAGE_CODES:
        flash("Please choose a supported language.", "danger")
        return _render_form()

    normalized_coach_column = _normalized_mobile_expression(Coach.mobile_number)
    if (
        Coach.query.filter(normalized_coach_column == mobile_number)
        .order_by(Coach.id.asc())
        .first()
    ):
        flash("This mobile number is already registered to a coach or administrator.", "danger")
        return _render_form()

    normalized_student_column = _normalized_mobile_expression(Student.mobile_number)
    if (
        Student.query.filter(normalized_student_column == mobile_number)
        .order_by(Student.id.asc())
        .first()
    ):
        flash("This mobile number is already registered to a student.", "danger")
        return _render_form()

    if email and Student.query.filter(Student.email == email).first():
        flash("This email is already registered to a student.", "danger")
        return _render_form()

    student = Student(
        name=name,
        mobile_number=mobile_number,
        email=email,
        state=state_choice,
        preferred_language=preferred_language,
    )
    student.set_password(password)
    db.session.add(student)

    summary: str | None = None
    rule_warning: str | None = None
    rule_exists = ExamRule.query.filter_by(state=state_choice).first() is not None
    try:
        db.session.flush()
        if rule_exists:
            summary = switch_student_state(
                student, state_choice, acting_student=student
            )
        else:
            db.session.commit()
            rule_warning = (
                "Exam rules for "
                f"{state_choice} are not configured yet."
                " Students can practise immediately, but administrators "
                "must add the rule before scheduling timed exams."
            )
    except (IntegrityError, StateSwitchError) as exc:
        db.session.rollback()
        if isinstance(exc, StateSwitchError):
            flash(str(exc), "danger")
        else:
            flash(
                "Unable to register with the provided details. Please try again.",
                "danger",
            )
        return _render_form()
    except Exception:
        db.session.rollback()
        flash(
            "Unexpected error while registering. Please try again in a moment.",
            "danger",
        )
        return redirect(url_for("coach.login"))
    except Exception:
        db.session.rollback()
        flash(
            "Unexpected error while registering. Please try again in a moment.",
            "danger",
        )
        return redirect(url_for("coach.login"))

    login_user(student)
    flash("Student account created successfully!", "success")
    if summary:
        flash(summary, "info")
    if rule_warning:
        flash(rule_warning, "warning")
    return redirect(url_for("student.dashboard"))


@coach_bp.route("/logout")
@login_required
def logout():
    logout_user()
    session.pop("preferred_language", None)
    flash("You have been logged out.", "info")
    return redirect(url_for("coach.login"))


@coach_bp.route("/dashboard")
@login_required
def dashboard():
    slot_query = AvailabilitySlot.query.filter(
        AvailabilitySlot.start_time >= datetime.utcnow()
    ).order_by(AvailabilitySlot.start_time.asc())
    if current_user.is_admin:
        upcoming_slots = slot_query.limit(5).all()
        student_count = Student.query.count()
        pending_bookings = (
            Appointment.query.filter(
                Appointment.status.in_(["booked", "pending_cancel"])
            ).count()
        )
    else:
        upcoming_slots = (
            slot_query.filter(AvailabilitySlot.coach_id == current_user.id)
            .limit(5)
            .all()
        )
        student_count = Student.query.filter_by(
            assigned_coach_id=current_user.id
        ).count()
        pending_bookings = (
            Appointment.query.join(AvailabilitySlot)
            .filter(AvailabilitySlot.coach_id == current_user.id)
            .filter(Appointment.status.in_(["booked", "pending_cancel"]))
            .count()
        )
    return render_template(
        "coach/dashboard.html",
        upcoming_slots=upcoming_slots,
        student_count=student_count,
        pending_bookings=pending_bookings,
    )


@coach_bp.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    if request.method == "POST":
        current_user.name = request.form.get("name", current_user.name)
        submitted_mobile_raw = (request.form.get("mobile_number") or "").strip()
        normalized_mobile = _normalize_mobile_number(submitted_mobile_raw)
        if not normalized_mobile:
            flash("Mobile number is required.", "warning")
            return render_template("coach/profile.html", state_choices=STATE_CHOICES)

        normalized_column = _normalized_mobile_expression(Coach.mobile_number)
        duplicate_mobile = (
            Coach.query.filter(normalized_column == normalized_mobile)
            .filter(Coach.id != current_user.id)
            .first()
        )
        if duplicate_mobile:
            flash("Another account already uses that mobile number.", "danger")
            return render_template("coach/profile.html", state_choices=STATE_CHOICES)

        current_user.mobile_number = normalized_mobile
        current_user.city = request.form.get("city", current_user.city)
        state_choice = (request.form.get("state") or "").strip().upper()
        if state_choice not in STATE_CHOICES:
            flash("Please choose a valid state or territory.", "warning")
            return render_template("coach/profile.html", state_choices=STATE_CHOICES)
        current_user.state = state_choice
        vehicle_type = _extract_vehicle_type_from_form()
        if not vehicle_type:
            flash("Please choose either automatic or manual transmission.", "warning")
            return render_template("coach/profile.html", state_choices=STATE_CHOICES)
        current_user.vehicle_types = vehicle_type
        current_user.bio = request.form.get("bio", current_user.bio)
        db.session.commit()
        flash("Profile updated successfully", "success")
        return redirect(url_for("coach.profile"))
    return render_template("coach/profile.html", state_choices=STATE_CHOICES)


@coach_bp.route("/students")
@login_required
def students():
    if current_user.is_admin:
        students = Student.query.order_by(Student.name.asc()).all()
    else:
        students = (
            Student.query.filter_by(assigned_coach_id=current_user.id)
            .order_by(Student.name.asc())
            .all()
        )
    summaries = {
        student.id: {
            "attempts": len(student.mock_exam_summaries),
            "last_score": student.mock_exam_summaries[-1].score
            if student.mock_exam_summaries
            else None,
        }
        for student in students
    }
    coach_lookup = {}
    if current_user.is_admin:
        coach_lookup = {coach.id: coach for coach in Coach.query.order_by(Coach.name).all()}
    return render_template(
        "coach/students.html",
        students=students,
        summaries=summaries,
        coach_lookup=coach_lookup,
    )


@coach_bp.route("/slots", methods=["GET", "POST"])
@login_required
def slots():
    if request.method == "POST":
        if current_user.is_admin:
            try:
                selected_coach_id = int(request.form.get("coach_id", ""))
            except (TypeError, ValueError):
                flash("Please choose a coach for the new slot.", "warning")
                return redirect(url_for("coach.slots"))
            if not db.session.get(Coach, selected_coach_id):
                flash("Selected coach could not be found.", "danger")
                return redirect(url_for("coach.slots"))
        else:
            selected_coach_id = current_user.id
        try:
            start_time = datetime.fromisoformat(request.form["start_time"])  # type: ignore[arg-type]
        except (KeyError, ValueError):
            flash("Invalid start time format", "danger")
            return redirect(url_for("coach.slots"))
        duration = int(request.form.get("duration", 30))
        if duration not in {30, 60}:
            flash("Duration must be either 30 or 60 minutes.", "warning")
            return redirect(url_for("coach.slots"))
        location_text = request.form.get("location", "").strip()
        if not location_text:
            flash("Location is required", "warning")
            return redirect(url_for("coach.slots"))
        slot = AvailabilitySlot(
            coach_id=selected_coach_id,
            start_time=start_time,
            duration_minutes=duration,
            location_text=location_text,
        )
        db.session.add(slot)
        try:
            db.session.commit()
            flash("Slot created", "success")
        except Exception as exc:  # broad to catch unique constraint
            db.session.rollback()
            flash("Unable to create slot: duplicate or invalid data", "danger")
        return redirect(url_for("coach.slots"))

    slot_query = AvailabilitySlot.query.order_by(AvailabilitySlot.start_time.asc())
    if not current_user.is_admin:
        slot_query = slot_query.filter_by(coach_id=current_user.id)
    slots = slot_query.all()
    coach_choices = []
    if current_user.is_admin:
        coach_choices = Coach.query.order_by(Coach.name.asc()).all()
    return render_template("coach/slots.html", slots=slots, coach_choices=coach_choices)


@coach_bp.route("/slots/<int:slot_id>/delete", methods=["POST"])
@login_required
def delete_slot(slot_id: int):
    slot_query = AvailabilitySlot.query.filter_by(id=slot_id)
    if not current_user.is_admin:
        slot_query = slot_query.filter_by(coach_id=current_user.id)
    slot = slot_query.first_or_404()
    if slot.appointment and slot.appointment.status == "booked":
        flash("Cannot delete a slot with an active booking.", "danger")
        return redirect(url_for("coach.slots"))
    db.session.delete(slot)
    db.session.commit()
    flash("Slot removed", "info")
    return redirect(url_for("coach.slots"))


@coach_bp.route("/appointments")
@login_required
def appointments():
    appointment_query = Appointment.query.join(AvailabilitySlot).order_by(
        AvailabilitySlot.start_time.desc()
    )
    if not current_user.is_admin:
        appointment_query = appointment_query.filter(
            AvailabilitySlot.coach_id == current_user.id
        )
    appointments = appointment_query.all()
    coach_lookup = {}
    if current_user.is_admin:
        coach_lookup = {coach.id: coach for coach in Coach.query.order_by(Coach.name).all()}
    return render_template(
        "coach/appointments.html",
        appointments=appointments,
        coach_lookup=coach_lookup,
    )


@coach_bp.route("/appointments/<int:appointment_id>/status", methods=["POST"])
@login_required
def update_appointment_status(appointment_id: int):
    appointment_query = Appointment.query.join(AvailabilitySlot)
    if not current_user.is_admin:
        appointment_query = appointment_query.filter(
            AvailabilitySlot.coach_id == current_user.id
        )
    appointment = (
        appointment_query.filter(Appointment.id == appointment_id).first_or_404()
    )
    status = request.form.get("status")
    valid_statuses = {"booked", "pending_cancel", "cancelled", "completed"}
    if status not in valid_statuses:
        flash("Invalid status", "danger")
        return redirect(url_for("coach.appointments"))
    appointment.status = status
    if status == "cancelled":
        appointment.slot.status = "available"
    elif status == "completed":
        appointment.slot.status = "unavailable"
    else:
        appointment.slot.status = "booked"
    if status == "booked":
        appointment.cancellation_requested_at = None
    elif status == "pending_cancel" and not appointment.cancellation_requested_at:
        appointment.cancellation_requested_at = datetime.utcnow()
    db.session.commit()
    flash("Appointment updated", "success")
    return redirect(url_for("coach.appointments"))


@coach_bp.route("/personnel", methods=["GET", "POST"])
@login_required
def personnel():
    redirect_response = _require_admin_access()
    if redirect_response:
        return redirect_response

    if request.method == "POST":
        form_type = request.form.get("form_type")
        if form_type == "create":
            _handle_account_creation()
        elif form_type == "update_password":
            _handle_password_update()
        else:
            flash("Unknown action requested.", "danger")
        return redirect(url_for("coach.personnel"))

    coaches = Coach.query.order_by(Coach.name.asc()).all()
    students = Student.query.order_by(Student.name.asc()).all()
    return render_template(
        "coach/personnel.html",
        coaches=coaches,
        students=students,
        coach_choices=coaches,
        state_choices=STATE_CHOICES,
    )


def _handle_account_creation() -> None:
    role = (request.form.get("role") or "").strip().lower()
    if role not in {"coach", "student", "admin"}:
        flash("Please choose a valid account type.", "warning")
        return

    if role in {"coach", "admin"}:
        name = (request.form.get("name") or "").strip()
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        mobile_input = (request.form.get("mobile_number") or "").strip()
        mobile_number = _normalize_mobile_number(mobile_input)
        city = (request.form.get("city") or "").strip()
        state = (request.form.get("state") or "").strip().upper()
        if state not in STATE_CHOICES:
            flash("Please choose a valid state or territory.", "warning")
            return
        vehicle_type = _extract_vehicle_type_from_form()

        if not all(
            [name, email, password, mobile_number, city, state, vehicle_type]
        ):
            flash(
                "All coach/admin fields are required, including a mobile number.",
                "warning",
            )
            return

        normalized_column = _normalized_mobile_expression(Coach.mobile_number)
        duplicate_mobile = (
            Coach.query.filter(normalized_column == mobile_number)
            .first()
        )
        if duplicate_mobile:
            flash("A coach or administrator already uses that mobile number.", "danger")
            return

        coach = Coach(
            name=name,
            email=email,
            mobile_number=mobile_number,
            city=city,
            state=state,
            vehicle_types=vehicle_type,
        )
        coach.set_password(password)
        db.session.add(coach)

        try:
            db.session.flush()
        except IntegrityError:
            db.session.rollback()
            flash(
                "Unable to create account: duplicate email or mobile number.",
                "danger",
            )
            return

        if role == "admin":
            db.session.add(Admin(id=coach.id))

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("Unable to create account due to duplicate information.", "danger")
            return

        flash("Account created successfully.", "success")
        return

    # Student creation
    name = (request.form.get("name") or "").strip()
    email = (request.form.get("email") or "").strip().lower()
    password = request.form.get("password") or ""
    mobile_input = (request.form.get("mobile_number") or "").strip()
    mobile_number = _normalize_mobile_number(mobile_input)
    state = (request.form.get("state") or "").strip().upper()
    if state not in STATE_CHOICES:
        flash("Please choose a valid state or territory.", "warning")
        return
    assigned_coach_raw = request.form.get("assigned_coach_id")
    assigned_coach = None
    if assigned_coach_raw:
        try:
            assigned_coach = db.session.get(Coach, int(assigned_coach_raw))
        except (TypeError, ValueError):
            assigned_coach = None

    if not all([name, email, password, mobile_number, state]):
        flash("All student fields are required.", "warning")
        return

    normalized_student_column = _normalized_mobile_expression(Student.mobile_number)
    duplicate_student = (
        Student.query.filter(normalized_student_column == mobile_number)
        .first()
    )
    if duplicate_student:
        flash("Another student already uses that mobile number.", "danger")
        return

    student = Student(
        name=name,
        email=email,
        mobile_number=mobile_number,
        state=state,
        preferred_language="ENGLISH",
    )
    if assigned_coach:
        student.coach = assigned_coach
    student.set_password(password)
    db.session.add(student)

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        flash("Unable to create student: duplicate email or mobile number.", "danger")
        return

    flash("Account created successfully.", "success")


def _handle_password_update() -> None:
    account_type = (request.form.get("account_type") or "").strip().lower()
    account_id = request.form.get("account_id")
    new_password = request.form.get("new_password") or ""

    if account_type not in {"coach", "student", "admin"}:
        flash("Unsupported account type for password update.", "danger")
        return

    try:
        identity = int(account_id)
    except (TypeError, ValueError):
        flash("Invalid account identifier.", "danger")
        return

    if len(new_password) < 6:
        flash("Please supply a password of at least 6 characters.", "warning")
        return

    if account_type == "student":
        entity = db.session.get(Student, identity)
        if not entity:
            flash("Student account not found.", "danger")
            return
        entity.set_password(new_password)
    else:
        coach = db.session.get(Coach, identity)
        if not coach:
            flash("Coach account not found.", "danger")
            return
        if account_type == "admin" and not coach.is_admin:
            flash("Selected account is not an administrator.", "danger")
            return
        coach.set_password(new_password)

    db.session.commit()
    flash("Password updated successfully.", "success")


def _handle_question_upload_action() -> None:
    file_storage = request.files.get("excel_file")
    if not file_storage or not file_storage.filename:
        flash("Please choose an Excel file to upload.", "warning")
        return

    default_language = (request.form.get("default_language") or "ENGLISH").strip().upper()
    if default_language not in LANGUAGE_CODES:
        default_language = "ENGLISH"

    selected_state = _default_question_state(request.form.get("default_state"))
    if selected_state not in STATE_CHOICES and selected_state != "ALL":
        flash("Please choose a valid default state for uploaded questions.", "warning")
        return

    try:
        file_storage.stream.seek(0)
        created, updated = _import_question_bank(
            file_storage.stream, default_state=selected_state, default_language=default_language
        )
    except ValueError as exc:
        flash(str(exc), "danger")
        return

    flash(f"Imported {created} new questions and updated {updated} existing records.", "success")


def _handle_exam_creation_action() -> None:
    title = (request.form.get("title") or "").strip()
    try:
        time_limit = int(request.form.get("time_limit", "0"))
    except ValueError:
        time_limit = 0
    if not title or time_limit <= 0:
        flash("Please provide a title and a positive time limit.", "warning")
        return

    selection_mode = (request.form.get("selection_mode") or "manual").strip().lower()
    state = _default_question_state(request.form.get("paper_state"))
    if state not in STATE_CHOICES:
        flash("Please choose a valid state for the exam paper.", "warning")
        return

    if not current_user.is_admin and state != current_user.state:
        state = current_user.state

    selected_questions: list[Question] = []
    if selection_mode == "manual":
        question_ids: list[int] = []
        for raw_id in request.form.getlist("question_ids"):
            try:
                question_ids.append(int(raw_id))
            except (TypeError, ValueError):
                continue
        if not question_ids:
            flash("Select at least one question for the paper.", "warning")
            return
        questions = Question.query.filter(Question.id.in_(question_ids)).all()
        lookup = {question.id: question for question in questions}
        for qid in question_ids:
            question = lookup.get(qid)
            if not question:
                continue
            if question.state_scope not in {"ALL", state}:
                continue
            selected_questions.append(question)
        if not selected_questions:
            flash("Selected questions do not match the chosen state.", "warning")
            return
    elif selection_mode == "auto":
        try:
            count = int(request.form.get("auto_count", "10"))
        except ValueError:
            count = 10
        count = max(1, min(count, 50))
        topic_filter = (request.form.get("auto_topic") or "").strip()
        query = _question_base_query(state if state != "ALL" else None)
        if topic_filter:
            query = query.filter(Question.topic.ilike(f"%{topic_filter}%"))
        pool = query.all()
        if not pool:
            flash("No questions available for the selected filters.", "warning")
            return
        if len(pool) <= count:
            selected_questions = pool
        else:
            selected_questions = random.sample(pool, count)
    else:
        flash("Unknown exam creation mode.", "danger")
        return

    paper = MockExamPaper(state=state, title=title, time_limit_minutes=time_limit)
    db.session.add(paper)
    db.session.flush()
    for position, question in enumerate(selected_questions, start=1):
        db.session.add(
            MockExamPaperQuestion(
                paper_id=paper.id,
                question_id=question.id,
                position=position,
            )
        )
    db.session.commit()
    flash("Exam paper created successfully.", "success")


def _handle_exam_delete_action() -> None:
    try:
        paper_id = int(request.form.get("paper_id", ""))
    except (TypeError, ValueError):
        flash("Invalid exam identifier.", "warning")
        return
    paper = db.session.get(MockExamPaper, paper_id)
    if not paper:
        flash("Exam paper not found.", "danger")
        return
    if not current_user.is_admin and paper.state != current_user.state:
        flash("You do not have permission to remove this paper.", "danger")
        return
    db.session.delete(paper)
    db.session.commit()
    flash("Exam paper removed.", "info")


@coach_bp.route("/exams", methods=["GET", "POST"])
@login_required
def exams():
    if request.method == "POST":
        action = request.form.get("action")
        if action == "upload_questions":
            _handle_question_upload_action()
        elif action == "create_exam":
            _handle_exam_creation_action()
        elif action == "delete_exam":
            _handle_exam_delete_action()
        else:
            flash("Unknown action.", "danger")
        return redirect(url_for("coach.exams"))

    selected_state = request.args.get("state") if current_user.is_admin else current_user.state
    if not selected_state:
        selected_state = current_user.state if not current_user.is_admin else "ALL"
    if selected_state not in STATE_CHOICES and selected_state != "ALL":
        selected_state = current_user.state if not current_user.is_admin else "ALL"

    paper_query = MockExamPaper.query.order_by(MockExamPaper.id.desc())
    if selected_state != "ALL":
        paper_query = paper_query.filter_by(state=selected_state)
    elif not current_user.is_admin:
        paper_query = paper_query.filter_by(state=current_user.state)
    papers = paper_query.all()

    question_state = None if selected_state == "ALL" else selected_state
    available_questions = _question_base_query(question_state).limit(300).all()

    return render_template(
        "coach/exams.html",
        papers=papers,
        available_questions=available_questions,
        state_choices=STATE_CHOICES,
        selected_state=selected_state,
        language_codes=LANGUAGE_CODES,
    )
def _parse_upload_headers(header_row: tuple) -> dict[int, str]:
    header_mapping = {
        "QID": "qid",
        "题目编号": "qid",
        "PROMPT": "prompt",
        "题干": "prompt",
        "OPTION A": "option_a",
        "选项A": "option_a",
        "OPTION B": "option_b",
        "选项B": "option_b",
        "OPTION C": "option_c",
        "选项C": "option_c",
        "OPTION D": "option_d",
        "选项D": "option_d",
        "CORRECT OPTION": "correct_option",
        "答案": "correct_option",
        "TOPIC": "topic",
        "考点类型": "topic",
        "EXPLANATION": "explanation",
        "解析": "explanation",
        "STATE SCOPE": "state_scope",
        "适用州": "state_scope",
        "LANGUAGE": "language",
        "语言": "language",
        "IMAGE URL": "image_url",
        "配图": "image_url",
    }
    column_map: dict[int, str] = {}
    for index, value in enumerate(header_row):
        if not value:
            continue
        label = str(value).strip().upper()
        field = header_mapping.get(label)
        if field:
            column_map[index] = field
    return column_map


def _import_question_bank(file_stream, *, default_state: str, default_language: str) -> tuple[int, int]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise ValueError(
            "Excel support requires the 'openpyxl' package. Install it with 'pip install openpyxl'."
        ) from exc

    try:
        workbook = load_workbook(file_stream, data_only=True)
    except Exception as exc:  # pragma: no cover - defensive
        raise ValueError(f"Unable to read Excel file: {exc}")

    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("Upload is missing a header row.")

    column_map = _parse_upload_headers(header_row)
    required_fields = {"prompt", "option_a", "option_b", "option_c", "option_d", "correct_option"}
    if not required_fields.issubset(set(column_map.values())):
        raise ValueError("Excel header must include prompt, options, and correct option columns.")

    created = 0
    updated = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        record: dict[str, str] = {}
        for index, value in enumerate(row):
            field = column_map.get(index)
            if not field or value is None:
                continue
            record[field] = str(value).strip()

        prompt = record.get("prompt")
        correct_option = (record.get("correct_option") or "").strip().upper()
        if not prompt or correct_option not in VALID_OPTIONS:
            continue

        qid = record.get("qid") or f"EXCEL-{uuid4().hex[:10].upper()}"
        topic = record.get("topic") or "general"
        explanation = record.get("explanation") or ""
        state_scope = (record.get("state_scope") or default_state or "ALL").upper()
        language = (record.get("language") or default_language or "ENGLISH").upper()
        image_url = record.get("image_url") or None

        option_a = record.get("option_a") or "Option A"
        option_b = record.get("option_b") or "Option B"
        option_c = record.get("option_c") or "Option C"
        option_d = record.get("option_d") or "Option D"

        question = Question.query.filter_by(qid=qid, state_scope=state_scope, language=language).first()
        if not question:
            question = Question(
                qid=qid,
                prompt=prompt,
                state_scope=state_scope,
                language=language,
                topic=topic,
                option_a=option_a,
                option_b=option_b,
                option_c=option_c,
                option_d=option_d,
                correct_option=correct_option,
                explanation=explanation,
                image_url=image_url,
            )
            db.session.add(question)
            created += 1
        else:
            question.prompt = prompt
            question.topic = topic
            question.option_a = option_a
            question.option_b = option_b
            question.option_c = option_c
            question.option_d = option_d
            question.correct_option = correct_option
            question.explanation = explanation
            question.image_url = image_url
            updated += 1

    db.session.commit()
    return created, updated
